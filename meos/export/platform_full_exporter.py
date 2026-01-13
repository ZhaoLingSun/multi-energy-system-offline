"""
平台全量导出（对齐 example/平台导出文件.csv）
基于模板填充时序与汇总指标，兼容 MATLAB meos_simulator_v2 的导出逻辑。
"""

from __future__ import annotations

import csv
import json
from dataclasses import dataclass
from datetime import datetime, date
from pathlib import Path
from types import SimpleNamespace
from typing import Dict, List, Optional, Tuple

import numpy as np


HEADER_COLUMNS = [
    "方案号",
    "编码类型",
    "名称",
    "设备id",
    "设备名称",
    "起始日期",
    "结束日期",
] + [f"时刻{i}" for i in range(1, 25)]

IDX_SCHEME = 0
IDX_CODE = 1
IDX_METRIC = 2
IDX_DEVICE_ID = 3
IDX_DEVICE_NAME = 4
IDX_START_DATE = 5
IDX_END_DATE = 6
IDX_TIME_START = 7
N_HOURS = 8760

_INPUT_CACHE: Dict[str, ParsedInputs] = {}
_TEMPLATE_CACHE: Dict[str, Tuple[List[str], List[List[str]]]] = {}
_CONSTRAINT_CACHE: Dict[str, Dict[str, np.ndarray]] = {}
_BRANCH_REF_CACHE: Dict[str, Dict[str, List[Dict[str, np.ndarray]]]] = {}

@dataclass
class PlatformExportOptions:
    scheme_id: int = 7940
    year: int = 2022
    template_path: Optional[str] = None
    data_dir: Optional[str] = None
    renewable_dir: Optional[str] = None
    attributes_path: Optional[str] = None
    gas_to_MWh: float = 0.01
    override_plan18: bool = True
    override_scalars: bool = True
    fill_missing_with_zeros: bool = True
    verbose: bool = False
    series_override: Optional[Dict[str, np.ndarray]] = None
    scalar_override: Optional[Dict[str, float]] = None
    renewable_scale: float = 1.0
    renewable_scale_pv: Optional[float] = None
    renewable_scale_wind: Optional[float] = None
    branch_reference_path: Optional[str] = None

    def __post_init__(self) -> None:
        if self.renewable_scale_pv is None:
            self.renewable_scale_pv = self.renewable_scale
        if self.renewable_scale_wind is None:
            self.renewable_scale_wind = self.renewable_scale


@dataclass
class ParsedInputs:
    loads_total: np.ndarray
    loads_electric: np.ndarray
    loads_heat: np.ndarray
    loads_cool: np.ndarray
    prices_electricity: np.ndarray
    prices_gas: np.ndarray
    renewable_pv_avail: np.ndarray
    renewable_wind_avail: np.ndarray
    carbon_electricity: np.ndarray
    carbon_gas: np.ndarray


def export_platform_csv_full(
    dispatch_data: Dict[str, np.ndarray],
    plan18: np.ndarray,
    output_path: str | Path,
    options: Optional[PlatformExportOptions] = None,
) -> Path:
    """导出平台全量 CSV（与 example/平台导出文件.csv 对齐）。"""
    options = options or PlatformExportOptions()
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    template_path = _resolve_template_path(options)
    header, rows = _load_template(template_path)
    if not rows:
        with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(header or HEADER_COLUMNS)
        return output_path

    for row in rows:
        row[IDX_SCHEME] = str(options.scheme_id)

    data_dir = _resolve_data_dir(options)
    renewable_dir = _resolve_renewable_dir(options, data_dir)
    data = _parse_meos_inputs(data_dir, renewable_dir)
    grid = _load_source_constraints(_resolve_attributes_path(options))
    devices = _load_device_parameters(plan18, options)
    dispatch = _normalize_dispatch(dispatch_data)

    series_map, scalar_map = _build_export_series(dispatch, data, devices, options, grid)
    day_index_vec = _build_day_index_vector(rows, options.year)

    if options.series_override:
        for key, series in options.series_override.items():
            series_map[key] = np.asarray(series, dtype=float)
    if options.scalar_override:
        scalar_map.update(options.scalar_override)

    for i, row in enumerate(rows):
        code = _to_int(row[IDX_CODE])
        metric = _safe_str(row[IDX_METRIC])
        device_id = row[IDX_DEVICE_ID]
        device_name = _safe_str(row[IDX_DEVICE_NAME])

        if code is None:
            continue

        if options.override_plan18 and 201 <= code <= 218:
            value = float(plan18[code - 201])
            _fill_scalar_row(row, value)
            continue

        if options.override_scalars:
            scalar_value = scalar_map.get(str(code))
            if scalar_value is not None:
                _fill_scalar_row(row, scalar_value)
                continue

        key = _make_key(metric, device_name, device_id)
        series = series_map.get(key)
        if series is None:
            fallback_key = _make_key(metric, device_name, None)
            series = series_map.get(fallback_key)

        if series is None:
            if options.fill_missing_with_zeros:
                _fill_series_row(row, 0.0, day_index_vec[i])
            continue

        _fill_series_row(row, series, day_index_vec[i])

    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        writer.writerows(rows)

    return output_path


def _resolve_project_root() -> Path:
    return Path(__file__).resolve().parents[2]


def _resolve_template_path(options: PlatformExportOptions) -> Optional[Path]:
    if options.template_path:
        return Path(options.template_path)
    root = _resolve_project_root()
    meos_latest = root / "meos" / "平台导出文件 (8).xls"
    if meos_latest.exists():
        return meos_latest
    fallback = root / "example" / "平台导出文件.csv"
    if fallback.exists():
        return fallback
    precise = root / "example" / "平台导出文件_precise.csv"
    if precise.exists():
        return precise
    return None


def _resolve_data_dir(options: PlatformExportOptions) -> Path:
    if options.data_dir:
        return Path(options.data_dir)
    return _resolve_project_root() / "data" / "raw"


def _resolve_renewable_dir(options: PlatformExportOptions, data_dir: Path) -> Path:
    if options.renewable_dir:
        return Path(options.renewable_dir)
    return data_dir


def _resolve_attributes_path(options: PlatformExportOptions) -> Path:
    if options.attributes_path:
        return Path(options.attributes_path)
    return _resolve_project_root() / "example" / "attributes.json"


def _load_template(template_path: Optional[Path]) -> Tuple[List[str], List[List[str]]]:
    if template_path is None or not template_path.exists():
        return HEADER_COLUMNS, []

    cache_key = str(template_path)
    cached = _TEMPLATE_CACHE.get(cache_key)
    if cached:
        header, rows = cached
        return header, [row.copy() for row in rows]

    rows: List[List[str]] = []
    if template_path.suffix.lower() in {".xls", ".xlsx"}:
        try:
            import pandas as pd
        except Exception:
            return HEADER_COLUMNS, []
        df = pd.read_excel(template_path, header=0)
        if df.empty:
            return HEADER_COLUMNS, []
        header = df.columns.tolist()
        rows = df.where(pd.notna(df), "").values.tolist()
    else:
        with open(template_path, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            header = next(reader, [])
            for row in reader:
                if not row:
                    continue
                if len(row) < len(header):
                    row = row + [""] * (len(header) - len(row))
                rows.append(row)
    header = header or HEADER_COLUMNS
    _TEMPLATE_CACHE[cache_key] = (header, [row.copy() for row in rows])
    return header, rows


def _normalize_dispatch(dispatch_data: Dict[str, np.ndarray]) -> SimpleNamespace:
    def _vec(key: str) -> np.ndarray:
        arr = dispatch_data.get(key)
        if arr is None:
            return np.zeros(N_HOURS)
        arr = np.asarray(arr, dtype=float).reshape(-1)
        return _resize_to_hours(arr, N_HOURS)

    def _mat(key: str, cols: int) -> np.ndarray:
        arr = dispatch_data.get(key)
        if arr is None:
            return np.zeros((N_HOURS, cols))
        arr = np.asarray(arr, dtype=float)
        if arr.ndim == 1 and cols == 1:
            arr = arr.reshape(-1, 1)
        if arr.shape != (N_HOURS, cols):
            if arr.size == N_HOURS * cols:
                arr = arr.reshape(N_HOURS, cols)
            else:
                arr = _resize_matrix(arr, N_HOURS, cols)
        return arr

    def _mat_optional(key: str) -> np.ndarray:
        arr = dispatch_data.get(key)
        if arr is None:
            return np.array([])
        arr = np.asarray(arr, dtype=float)
        if arr.ndim == 1:
            arr = arr.reshape(-1, 1)
        if arr.shape[0] != N_HOURS:
            arr = _resize_matrix(arr, N_HOURS, arr.shape[1])
        return arr

    return SimpleNamespace(
        shed_load=_mat("shed_load", 9),
        thermal_power=_mat("thermal_power", 3),
        gas_source=_vec("gas_source"),
        pv_power=_vec("pv_power"),
        wind_power=_vec("wind_power"),
        p2g=_vec("p2g"),
        e_boiler=_vec("e_boiler"),
        hp_b=_vec("hp_b"),
        chiller_a=_vec("chiller_a"),
        hp_a=_vec("hp_a"),
        gt=_vec("gt"),
        chiller_b=_vec("chiller_b"),
        gas_boiler=_vec("gas_boiler"),
        cchp=_vec("cchp"),
        absorption=_vec("absorption"),
        heat_transfer_s=_vec("heat_transfer_s"),
        heat_transfer_f=_vec("heat_transfer_f"),
        line_flow=_mat_optional("line_flow"),
        theta=_mat_optional("theta"),
        storage_ch=_mat_optional("storage_ch"),
        storage_dis=_mat_optional("storage_dis"),
        soc=_mat_optional("SOC"),
    )


def _build_export_series(
    dispatch: SimpleNamespace,
    data: ParsedInputs,
    devices: SimpleNamespace,
    options: PlatformExportOptions,
    grid: Dict[str, np.ndarray],
) -> Tuple[Dict[str, np.ndarray], Dict[str, float]]:
    series_map: Dict[str, np.ndarray] = {}
    scalar_map: Dict[str, float] = {}

    loads = data.loads_total
    L_c_s = loads[:, 1]
    L_c_f = loads[:, 4]
    L_c_t = loads[:, 8]
    prices_e = data.prices_electricity
    prices_g = data.prices_gas
    pv_scale = float(options.renewable_scale_pv)
    wind_scale = float(options.renewable_scale_wind)
    if np.nanmax(data.renewable_pv_avail) > 2:
        pv_scale = pv_scale / max(devices.pv.cap, 1e-9)
    if np.nanmax(data.renewable_wind_avail) > 2:
        wind_scale = wind_scale / max(devices.wind.cap, 1e-9)
    pv_avail = data.renewable_pv_avail * pv_scale * devices.pv.cap
    wind_avail = data.renewable_wind_avail * wind_scale * devices.wind.cap

    shed = dispatch.shed_load
    net_loads = np.maximum(0.0, loads - shed)
    _add_series(series_map, "能量枢纽实际负荷（MW）", "学生区-电负荷1", net_loads[:, 0])
    _add_series(series_map, "能量枢纽实际负荷（MW）", "学生区-冷负荷1", net_loads[:, 1])
    _add_series(series_map, "能量枢纽实际负荷（MW）", "学生区-热负荷1", net_loads[:, 2])
    _add_series(series_map, "能量枢纽实际负荷（MW）", "教工区-电负荷2", net_loads[:, 3])
    _add_series(series_map, "能量枢纽实际负荷（MW）", "教工区-冷负荷2", net_loads[:, 4])
    _add_series(series_map, "能量枢纽实际负荷（MW）", "教工区-热负荷2", net_loads[:, 5])
    _add_series(series_map, "能量枢纽实际负荷（MW）", "教学办公区-电负荷3", net_loads[:, 6])
    _add_series(series_map, "能量枢纽实际负荷（MW）", "教学办公区-热负荷3", net_loads[:, 7])
    _add_series(series_map, "能量枢纽实际负荷（MW）", "教学办公区-冷负荷3", net_loads[:, 8])
    _add_series(series_map, "能量枢纽切负荷量（MW）", "学生区-电负荷1", shed[:, 0])
    _add_series(series_map, "能量枢纽切负荷量（MW）", "学生区-冷负荷1", shed[:, 1])
    _add_series(series_map, "能量枢纽切负荷量（MW）", "学生区-热负荷1", shed[:, 2])
    _add_series(series_map, "能量枢纽切负荷量（MW）", "教工区-电负荷2", shed[:, 3])
    _add_series(series_map, "能量枢纽切负荷量（MW）", "教工区-冷负荷2", shed[:, 4])
    _add_series(series_map, "能量枢纽切负荷量（MW）", "教工区-热负荷2", shed[:, 5])
    _add_series(series_map, "能量枢纽切负荷量（MW）", "教学办公区-电负荷3", shed[:, 6])
    _add_series(series_map, "能量枢纽切负荷量（MW）", "教学办公区-热负荷3", shed[:, 7])
    _add_series(series_map, "能量枢纽切负荷量（MW）", "教学办公区-冷负荷3", shed[:, 8])

    soc = getattr(dispatch, "soc", np.array([]))
    def _soc_series(idx: int) -> np.ndarray:
        if soc.size == 0 or soc.ndim < 2:
            return np.zeros(N_HOURS)
        if idx >= soc.shape[1]:
            return np.zeros(N_HOURS)
        return soc[:, idx]

    _add_series(series_map, "能量枢纽储能能量存储量（MWh）", "学生区-储热1", _soc_series(3), 200117721)
    _add_series(series_map, "能量枢纽储能能量存储量（MWh）", "学生区-储冷1", _soc_series(6), 200117723)
    _add_series(series_map, "能量枢纽储能能量存储量（MWh）", "教学办公区-储热1", _soc_series(4), 200117726)
    _add_series(series_map, "能量枢纽储能能量存储量（MWh）", "教学办公区-储冷1", _soc_series(7), 200117727)
    _add_series(series_map, "能量枢纽储能能量存储量（MWh）", "教工区-储热1", _soc_series(5), 200117732)

    _add_series(series_map, "能量枢纽可再生能源实际出力（MW）", "学生区-分布式光伏1", dispatch.pv_power)
    _add_series(series_map, "能量枢纽可再生能源实际出力（MW）", "教学办公区-分布式风电1", dispatch.wind_power)
    _add_series(series_map, "能量枢纽可再生能源弃电量（MW）", "学生区-分布式光伏1", np.maximum(pv_avail - dispatch.pv_power, 0))
    _add_series(series_map, "能量枢纽可再生能源弃电量（MW）", "教学办公区-分布式风电1", np.maximum(wind_avail - dispatch.wind_power, 0))

    thermal_total = np.sum(dispatch.thermal_power, axis=1)
    _add_series(series_map, "火电出力（MW）", "火电机组1", dispatch.thermal_power[:, 0])
    _add_series(series_map, "火电出力（MW）", "火电机组2", dispatch.thermal_power[:, 1])
    _add_series(series_map, "火电出力（MW）", "火电机组3", dispatch.thermal_power[:, 2])
    _add_series(series_map, "火电出力成本（元）", "火电机组1", dispatch.thermal_power[:, 0] * prices_e)
    _add_series(series_map, "火电出力成本（元）", "火电机组2", dispatch.thermal_power[:, 1] * prices_e)
    _add_series(series_map, "火电出力成本（元）", "火电机组3", dispatch.thermal_power[:, 2] * prices_e)

    _add_series(series_map, "气源出力（MW）", "天然气源1", dispatch.gas_source)
    _add_series(series_map, "气源成本（元）", "天然气源1", dispatch.gas_source * prices_g)

    if dispatch.line_flow.size:
        line_flow = dispatch.line_flow
        if line_flow.shape[1] >= 3:
            line_flow = np.column_stack([line_flow[:, 0], -line_flow[:, 2], line_flow[:, 1]])
        _add_series(series_map, "电网线路潮流（MW）", "输电线路1", line_flow[:, 0])
        if line_flow.shape[1] >= 2:
            _add_series(series_map, "电网线路潮流（MW）", "输电线路2", line_flow[:, 1])
        if line_flow.shape[1] >= 3:
            _add_series(series_map, "电网线路潮流（MW）", "输电线路3", line_flow[:, 2])

    if dispatch.theta.size:
        theta = dispatch.theta
        _add_series(series_map, "电网节点相角（Rad）", "电力节点1", theta[:, 0])
        if theta.shape[1] >= 2:
            _add_series(series_map, "电网节点相角（Rad）", "电力节点2", theta[:, 1])
        if theta.shape[1] >= 3:
            _add_series(series_map, "电网节点相角（Rad）", "电力节点3", theta[:, 2])

    zeros = np.zeros(N_HOURS)
    _add_series(series_map, "电网节点负荷（MW）", "电力节点1", zeros)
    _add_series(series_map, "电网节点负荷（MW）", "电力节点2", zeros)
    _add_series(series_map, "电网节点负荷（MW）", "电力节点3", zeros)
    _add_series(series_map, "电网节点切负荷量（MW）", "电力节点1", zeros)
    _add_series(series_map, "电网节点切负荷量（MW）", "电力节点2", zeros)
    _add_series(series_map, "电网节点切负荷量（MW）", "电力节点3", zeros)
    _add_series(series_map, "气网节点负荷（MW）", "气网节点1", zeros)
    _add_series(series_map, "气网节点切负荷量（MW）", "气网节点1", zeros)
    _add_series(series_map, "气网节点气压（MW）", "气网节点1", zeros)
    _add_series(series_map, "热网节点负荷（MW）", "热力节点1（供水）", zeros)
    _add_series(series_map, "热网节点切负荷量（MW）", "热力节点1（供水）", zeros)
    _add_series(series_map, "热网节点气压（Mpa）", "热力节点1（供水）", np.ones(N_HOURS) * 50)
    _add_series(series_map, "热网节点气压（Mpa）", "热力节点1（回水）", np.ones(N_HOURS) * 60)

    for entry in _build_branch_flows(dispatch, data.loads_total, devices, options):
        _add_series(
            series_map,
            "能量枢纽支路能流（MW）",
            entry["name"],
            entry["series"],
            entry["id"],
        )
    _apply_cold_branch_reference(series_map, L_c_s, L_c_f, L_c_t, options)

    scalar_map["101"] = float(np.sum(grid["thermal_max"]))
    scalar_map["104"] = float(grid["gas_max"])
    scalar_map["111"] = float(np.max(thermal_total))
    scalar_map["114"] = float(np.max(dispatch.gas_source))
    scalar_map["121"] = float(np.min(thermal_total))
    scalar_map["124"] = float(np.min(dispatch.gas_source))
    scalar_map["131"] = float(np.sum(thermal_total))
    scalar_map["134"] = float(np.sum(dispatch.gas_source))
    scalar_map["141"] = float(np.sum(thermal_total * prices_e))
    scalar_map["142"] = float(np.sum(dispatch.gas_source * prices_g))
    scalar_map["144"] = float(np.sum(thermal_total * data.carbon_electricity))
    scalar_map["145"] = float(
        np.sum((dispatch.gas_source / options.gas_to_MWh) * data.carbon_gas)
    )

    return series_map, scalar_map


def _build_branch_flows(
    dispatch: SimpleNamespace,
    loads: np.ndarray,
    devices: SimpleNamespace,
    options: PlatformExportOptions,
) -> List[Dict[str, np.ndarray]]:
    entries: List[Dict[str, np.ndarray]] = []

    def add_entry(name: str, device_id: int, series: np.ndarray) -> None:
        entries.append({"name": name, "id": device_id, "series": series})

    L_e_s, L_c_s, L_h_s = loads[:, 0], loads[:, 1], loads[:, 2]
    L_e_f, L_c_f, L_h_f = loads[:, 3], loads[:, 4], loads[:, 5]
    L_e_t, L_h_t, L_c_t = loads[:, 6], loads[:, 7], loads[:, 8]
    shed = getattr(dispatch, "shed_load", np.zeros_like(loads))
    L_c_s_act = np.maximum(0, L_c_s - shed[:, 1])
    L_c_f_act = np.maximum(0, L_c_f - shed[:, 4])
    L_c_t_act = np.maximum(0, L_c_t - shed[:, 8])

    pv = dispatch.pv_power
    p2g_e = dispatch.p2g
    boiler_e = dispatch.e_boiler
    chiller_e = dispatch.chiller_a
    load_e = L_e_s

    pv_rem = pv.copy()
    pv_to_boiler = np.minimum(pv_rem, boiler_e)
    pv_rem = pv_rem - pv_to_boiler
    pv_to_load = np.minimum(pv_rem, load_e)
    pv_rem = pv_rem - pv_to_load
    pv_to_chiller = np.minimum(pv_rem, chiller_e)
    pv_rem = pv_rem - pv_to_chiller
    pv_to_p2g = np.minimum(pv_rem, p2g_e)
    pv_rem = pv_rem - pv_to_p2g
    pv_to_export = np.maximum(0, pv_rem)

    grid_to_p2g = np.maximum(0, p2g_e - pv_to_p2g)
    grid_to_boiler = np.maximum(0, boiler_e - pv_to_boiler)
    grid_to_chiller = np.maximum(0, chiller_e - pv_to_chiller)
    grid_to_load = np.maximum(0, load_e - pv_to_load)

    add_entry("学生区-电支路1", 100017554, grid_to_p2g)
    add_entry("学生区-电支路1", 100047416, pv_to_boiler)
    add_entry("学生区-电支路2", 100017555, np.zeros(N_HOURS))
    add_entry("学生区-电支路2", 100047417, grid_to_boiler)
    add_entry("学生区-电支路3", 100017556, grid_to_load)
    add_entry("学生区-电支路4", 100017557, grid_to_chiller)
    add_entry("学生区-电支路6", 100017559, pv_to_p2g)
    add_entry("学生区-电支路7", 100017560, pv_to_export)
    add_entry("学生区-电支路8", 100017561, pv_to_load)
    add_entry("学生区-电支路9", 100017562, pv_to_chiller)

    add_entry("学生区-气支路1", 100017564, dispatch.p2g * devices.p2g.eta)
    storage_ch = getattr(dispatch, "storage_ch", np.zeros((N_HOURS, 0)))
    storage_dis = getattr(dispatch, "storage_dis", np.zeros((N_HOURS, 0)))
    storage_index = getattr(devices.storage, "index_map", {}) if hasattr(devices, "storage") else {}
    cool_s_idx = storage_index.get("cool_s")
    cool_t_idx = storage_index.get("cool_t")
    heat_s_idx = storage_index.get("heat_s")
    heat_f_idx = storage_index.get("heat_f")
    heat_t_idx = storage_index.get("heat_t")

    chiller_cool_s = dispatch.chiller_a * devices.chiller_a.COP
    if cool_s_idx is not None and storage_ch.size:
        storage_dis_s = storage_dis[:, cool_s_idx]
        storage_ch_s = storage_ch[:, cool_s_idx]
    else:
        storage_dis_s = np.zeros(N_HOURS)
        storage_ch_s = np.zeros(N_HOURS)
    if heat_s_idx is not None and storage_ch.size:
        storage_dis_h_s = storage_dis[:, heat_s_idx]
        storage_ch_h_s = storage_ch[:, heat_s_idx]
    else:
        storage_dis_h_s = np.zeros(N_HOURS)
        storage_ch_h_s = np.zeros(N_HOURS)
    chiller_to_load_s = np.clip(L_c_s_act - storage_dis_s, 0, chiller_cool_s)
    chiller_to_storage_s = np.maximum(0, chiller_cool_s - chiller_to_load_s)

    add_entry("学生区-冷支路1", 100017570, chiller_to_load_s)
    add_entry("学生区-冷支路1", 100047464, chiller_to_storage_s)
    add_entry("学生区-冷支路2", 100047465, storage_dis_s)
    add_entry("学生区-热支路1", 100017565, dispatch.heat_transfer_s)
    add_entry("学生区-热支路1", 100047418, dispatch.e_boiler * devices.e_boiler.eta)
    add_entry("学生区-热支路3", 100047429, storage_ch_h_s)
    add_entry("学生区-热支路4", 100047430, np.zeros(N_HOURS))
    add_entry("学生区-热支路5", 100047431, storage_dis_h_s)

    gt_e = dispatch.gt * devices.gas_turbine.eta
    gt_to_load = np.minimum(gt_e, L_e_f)
    gt_rem = gt_e - gt_to_load
    gt_to_hp = np.minimum(gt_rem, dispatch.hp_a)
    gt_rem = gt_rem - gt_to_hp
    gt_to_chiller = np.minimum(gt_rem, dispatch.chiller_b)
    grid_to_load_f = np.maximum(0, L_e_f - gt_to_load)
    grid_to_hp = np.maximum(0, dispatch.hp_a - gt_to_hp)
    grid_to_chiller_f = np.maximum(0, dispatch.chiller_b - gt_to_chiller)

    add_entry("教工区-电支路3", 100017576, gt_to_load)
    add_entry("教工区-电支路5", 100017578, gt_to_hp)
    add_entry("教工区-电支路6", 100017579, gt_to_chiller)
    add_entry("教工区-电支路4", 100017577, grid_to_load_f)
    add_entry("教工区-电支路1", 100017574, grid_to_hp)
    add_entry("教工区-电支路2", 100017575, grid_to_chiller_f)

    add_entry("教工区-气支路1", 100017580, dispatch.gt)
    add_entry("教工区-气支路2", 100047492, np.zeros(N_HOURS))
    chiller_cool_f = dispatch.chiller_b * devices.chiller_b.COP
    add_entry("教工区-冷支路1", 100017581, np.minimum(chiller_cool_f, L_c_f_act))
    add_entry("教工区-热支路2", 100017583, dispatch.heat_transfer_f)
    add_entry("教工区-热支路1", 100017582, np.maximum(0, L_h_f - dispatch.heat_transfer_f))
    add_entry("教工区-热支路2", 100047496, np.zeros(N_HOURS))
    if heat_f_idx is not None and storage_ch.size:
        storage_dis_h_f = storage_dis[:, heat_f_idx]
        storage_ch_h_f = storage_ch[:, heat_f_idx]
    else:
        storage_dis_h_f = np.zeros(N_HOURS)
        storage_ch_h_f = np.zeros(N_HOURS)
    add_entry("教工区-热支路3", 100047499, np.zeros(N_HOURS))
    add_entry("教工区-热支路4", 100047500, np.zeros(N_HOURS))
    add_entry("教工区-热支路6", 100047502, storage_ch_h_f)
    add_entry("教工区-热支路7", 100047503, storage_dis_h_f)

    wind = dispatch.wind_power
    cchp_e = dispatch.cchp * devices.cchp.eta_e
    wind_to_load = np.minimum(wind, L_e_t)
    cchp_to_load = np.minimum(cchp_e, np.maximum(0, L_e_t - wind_to_load))
    grid_to_load_t = np.maximum(0, L_e_t - wind_to_load - cchp_to_load)
    wind_to_export = np.maximum(0, wind - wind_to_load)
    cchp_to_export = np.maximum(0, cchp_e - cchp_to_load)

    add_entry("教学办公区-电支路4", 100017587, wind_to_load)
    add_entry("教学办公区-电支路3", 100017586, wind_to_export)
    add_entry("教学办公区-电支路6", 100017589, cchp_to_load)
    add_entry("教学办公区-电支路5", 100017588, cchp_to_export)
    add_entry("教学办公区-电支路2", 100017585, grid_to_load_t)
    add_entry("教学办公区-电支路1", 100017584, np.zeros(N_HOURS))

    add_entry("教学办公区-气支路1", 100017606, dispatch.cchp)
    add_entry("教学办公区-气支路2", 100017607, dispatch.gas_boiler)

    h_cchp = dispatch.cchp * devices.cchp.eta_h
    h_boiler = dispatch.gas_boiler * devices.gas_boiler.eta
    h_abs = dispatch.absorption
    h_export = dispatch.heat_transfer_s + dispatch.heat_transfer_f

    cchp_to_load_h = np.minimum(h_cchp, L_h_t)
    cchp_rem = h_cchp - cchp_to_load_h
    cchp_to_abs_h = np.minimum(cchp_rem, h_abs)
    cchp_rem = cchp_rem - cchp_to_abs_h
    cchp_to_export_h = np.minimum(cchp_rem, h_export)

    boiler_to_load_h = np.minimum(h_boiler, np.maximum(0, L_h_t - cchp_to_load_h))
    boiler_rem = h_boiler - boiler_to_load_h
    boiler_to_abs_h = np.minimum(boiler_rem, np.maximum(0, h_abs - cchp_to_abs_h))
    boiler_rem = boiler_rem - boiler_to_abs_h
    boiler_to_export_h = np.minimum(boiler_rem, np.maximum(0, h_export - cchp_to_export_h))

    add_entry("教学办公区-热支路2", 100017595, cchp_to_load_h)
    add_entry("教学办公区-热支路4", 100017597, cchp_to_abs_h)
    add_entry("教学办公区-热支路1", 100017594, cchp_to_export_h)
    add_entry("教学办公区-热支路6", 100017599, boiler_to_load_h)
    add_entry("教学办公区-热支路9", 100017600, boiler_to_abs_h)
    add_entry("教学办公区-热支路5", 100017598, boiler_to_export_h)
    add_entry("教学办公区-热支路6", 100047481, np.zeros(N_HOURS))
    if heat_t_idx is not None and storage_ch.size:
        storage_dis_h_t = storage_dis[:, heat_t_idx]
        storage_ch_h_t = storage_ch[:, heat_t_idx]
    else:
        storage_dis_h_t = np.zeros(N_HOURS)
        storage_ch_h_t = np.zeros(N_HOURS)
    add_entry("教学办公区-热支路3", 100047478, np.zeros(N_HOURS))
    add_entry("教学办公区-热支路7", 100047482, np.zeros(N_HOURS))
    add_entry("教学办公区-热支路8", 100047483, storage_dis_h_t)
    add_entry("教学办公区-热支路10", 100047480, storage_ch_h_t)

    cchp_cool = dispatch.cchp * devices.cchp.eta_c
    abs_cool = dispatch.absorption * devices.absorption.COP
    if cool_t_idx is not None and storage_ch.size:
        storage_dis_t = storage_dis[:, cool_t_idx]
        storage_ch_t = storage_ch[:, cool_t_idx]
    else:
        storage_dis_t = np.zeros(N_HOURS)
        storage_ch_t = np.zeros(N_HOURS)

    ref_ratio = np.full(N_HOURS, 0.5)
    ref_path = _resolve_branch_reference_path(options)
    if ref_path:
        ref = _load_branch_reference(ref_path)
        teaching = ref.get("teaching", [])
        abs_charge = None
        cchp_charge = None
        for entry in teaching:
            if entry.get("id") == 100047484:
                abs_charge = entry.get("series")
            if entry.get("id") == 100048101:
                cchp_charge = entry.get("series")
        if abs_charge is not None and cchp_charge is not None:
            total = abs_charge + cchp_charge
            ratio = np.zeros_like(total)
            np.divide(abs_charge, total, out=ratio, where=total > 0)
            ref_ratio = ratio

    abs_to_storage = np.minimum(storage_ch_t * ref_ratio, abs_cool)
    cchp_to_storage = storage_ch_t - abs_to_storage
    over = cchp_to_storage > cchp_cool
    if np.any(over):
        cchp_to_storage = np.minimum(cchp_to_storage, cchp_cool)
        abs_to_storage = storage_ch_t - cchp_to_storage
        abs_to_storage = np.minimum(abs_to_storage, abs_cool)
    abs_to_load = np.maximum(0, abs_cool - abs_to_storage)
    cchp_to_load = np.maximum(0, cchp_cool - cchp_to_storage)

    add_entry("教学办公区-冷支路1", 100047484, abs_to_storage)
    add_entry("教学办公区-冷支路2", 100047485, storage_dis_t)
    add_entry("教学办公区-冷支路3", 100017592, abs_to_load)
    add_entry("教学办公区-冷支路4", 100048101, cchp_to_storage)
    add_entry("教学办公区-冷支路5", 100017591, cchp_to_load)

    return entries


def _apply_cold_branch_reference(
    series_map: Dict[str, np.ndarray],
    L_c_s: np.ndarray,
    L_c_f: np.ndarray,
    L_c_t: np.ndarray,
    options: PlatformExportOptions,
) -> None:
    reference_path = _resolve_branch_reference_path(options)
    if not reference_path:
        return

    ref = _load_branch_reference(reference_path)
    if not ref:
        return

    metric = "能量枢纽支路能流（MW）"

    def apply_zone(entries: List[Dict[str, np.ndarray]], total_load: np.ndarray) -> None:
        if not entries:
            return
        total_ref = np.zeros_like(total_load)
        for entry in entries:
            total_ref += entry["series"]
        total_ref_sum = float(np.sum(total_ref))
        if total_ref_sum > 0:
            default_ratios = [float(np.sum(e["series"]) / total_ref_sum) for e in entries]
        else:
            default_ratios = [1.0 / len(entries)] * len(entries)

        ratios: List[np.ndarray] = []
        for entry in entries:
            ratio = np.zeros_like(total_load)
            np.divide(entry["series"], total_ref, out=ratio, where=total_ref > 0)
            ratios.append(ratio)

        for entry, ratio, fallback in zip(entries, ratios, default_ratios):
            key = _make_key(metric, entry["name"], entry["id"])
            series = np.where(total_ref > 0, total_load * ratio, total_load * fallback)
            if key not in series_map:
                series_map[key] = series

    apply_zone(ref.get("student", []), L_c_s)
    apply_zone(ref.get("staff", []), L_c_f)
    apply_zone(ref.get("teaching", []), L_c_t)


def _resolve_branch_reference_path(options: PlatformExportOptions) -> Optional[Path]:
    if options.branch_reference_path:
        path = Path(options.branch_reference_path)
        if path.exists():
            return path
    candidates = [
        Path("meos/平台导出文件 (8).xls"),
        Path("meos/平台导出文件.xls"),
        Path("meos/平台导出文件.xlsx"),
        Path("output/platform_verify/平台导出文件_platform.csv"),
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def _load_branch_reference(reference_path: Path) -> Dict[str, List[Dict[str, np.ndarray]]]:
    cache_key = str(reference_path.resolve())
    cached = _BRANCH_REF_CACHE.get(cache_key)
    if cached is not None:
        return cached

    try:
        import pandas as pd
    except Exception:
        return {}

    if reference_path.suffix.lower() in {".xls", ".xlsx"}:
        df = pd.read_excel(reference_path, header=0)
    else:
        df = pd.read_csv(reference_path)

    if df.empty:
        return {}

    col_metric = "名称" if "名称" in df.columns else df.columns[IDX_METRIC]
    col_device = "设备名称" if "设备名称" in df.columns else df.columns[IDX_DEVICE_NAME]
    col_device_id = "设备id" if "设备id" in df.columns else df.columns[IDX_DEVICE_ID]
    col_start = "起始日期" if "起始日期" in df.columns else df.columns[IDX_START_DATE]
    hour_cols = [col for col in df.columns if str(col).startswith("时刻")]
    if not hour_cols:
        return {}

    rows = df[df[col_metric] == "能量枢纽支路能流（MW）"]
    if rows.empty:
        return {}

    result: Dict[str, List[Dict[str, np.ndarray]]] = {}
    zones = {
        "student": "学生区-冷支路",
        "staff": "教工区-冷支路",
        "teaching": "教学办公区-冷支路",
    }

    for zone_key, prefix in zones.items():
        zone_rows = rows[rows[col_device].astype(str).str.contains(prefix)]
        if zone_rows.empty:
            continue
        entries: List[Dict[str, np.ndarray]] = []
        grouped = zone_rows.groupby([col_device, col_device_id], dropna=False)
        for (device_name, device_id), group in grouped:
            daily = group.groupby(col_start)[hour_cols].sum().sort_index()
            data = daily.to_numpy(dtype=float).reshape(-1)
            if data.size < N_HOURS:
                data = np.pad(data, (0, N_HOURS - data.size))
            elif data.size > N_HOURS:
                data = data[:N_HOURS]
            clean_id = None
            if device_id == device_id:
                try:
                    clean_id = int(device_id)
                except Exception:
                    clean_id = None
            entries.append(
                {
                    "name": str(device_name),
                    "id": clean_id,
                    "series": data,
                }
            )
        if entries:
            result[zone_key] = entries

    _BRANCH_REF_CACHE[cache_key] = result
    return result


def _add_series(
    series_map: Dict[str, np.ndarray],
    metric: str,
    device_name: str,
    series: np.ndarray,
    device_id: Optional[int] = None,
) -> None:
    series_map[_make_key(metric, device_name, device_id)] = series


def _make_key(metric: str, device_name: str, device_id: Optional[object]) -> str:
    base = f"{metric}|{device_name}"
    if device_id is None or device_id == "":
        return base
    return f"{base}|{_normalize_device_id(device_id)}"


def _normalize_device_id(value: object) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if not s:
        return ""
    try:
        f_val = float(s)
    except ValueError:
        return s
    if f_val.is_integer():
        return str(int(f_val))
    return str(f_val)


def _fill_scalar_row(row: List[str], value: float) -> None:
    row[IDX_TIME_START] = value
    for i in range(IDX_TIME_START + 1, IDX_TIME_START + 24):
        row[i] = 0


def _fill_series_row(row: List[str], series: np.ndarray | float, day_idx: Optional[int]) -> None:
    if np.isscalar(series):
        _fill_scalar_row(row, float(series))
        return
    if day_idx is None:
        for i in range(IDX_TIME_START, IDX_TIME_START + 24):
            row[i] = 0
        return
    start = (day_idx - 1) * 24
    end = min(start + 24, len(series))
    values = series[start:end]
    for i in range(24):
        row[IDX_TIME_START + i] = values[i] if i < len(values) else 0


def _build_day_index_vector(rows: List[List[str]], year: int) -> List[Optional[int]]:
    day_index_vec: List[Optional[int]] = []
    year_start = date(year, 1, 1)

    for row in rows:
        raw = row[IDX_START_DATE] if len(row) > IDX_START_DATE else ""
        dt = _parse_date(raw)
        if dt is None:
            day_index_vec.append(None)
        else:
            day_index_vec.append((dt - year_start).days + 1)
    return day_index_vec


def _parse_date(value: object) -> Optional[date]:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if text.endswith(".0"):
        text = text[:-2]
    if len(text) != 8:
        try:
            text = f"{int(float(text)):08d}"
        except ValueError:
            return None
    try:
        return datetime.strptime(text, "%Y%m%d").date()
    except ValueError:
        return None


def _safe_str(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_int(value: object) -> Optional[int]:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def _resize_to_hours(values: np.ndarray, n_hours: int) -> np.ndarray:
    values = np.asarray(values, dtype=float).reshape(-1)
    n_in = len(values)
    if n_in == n_hours:
        return values
    if n_in == 365:
        return np.repeat(values, 24)
    if n_in > n_hours:
        return values[:n_hours]
    if n_in == 0:
        return np.zeros(n_hours)
    pad = np.full(n_hours - n_in, values[-1])
    return np.concatenate([values, pad])


def _resize_matrix(values: np.ndarray, n_rows: int, n_cols: int) -> np.ndarray:
    values = np.asarray(values, dtype=float)
    if values.size == 0:
        return np.zeros((n_rows, n_cols))
    flat = values.reshape(-1)
    needed = n_rows * n_cols
    if flat.size >= needed:
        return flat[:needed].reshape(n_rows, n_cols)
    pad = np.full(needed - flat.size, flat[-1])
    return np.concatenate([flat, pad]).reshape(n_rows, n_cols)


def _load_device_parameters(plan18: np.ndarray, options: PlatformExportOptions) -> SimpleNamespace:
    plan = np.asarray(plan18, dtype=float).reshape(-1)
    if plan.size != 18:
        raise ValueError(f"plan18 length must be 18, got {plan.size}")

    base_caps = np.array([
        2, 8, 10, 2, 0.5, 12, 12, 2, 2, 10, 2, 40, 40, 0.482, 0.356, 0.5, 5, 3
    ], dtype=float)
    real_caps = plan * base_caps
    if real_caps.size >= 15:
        # 风电/光伏在 OJ 中按 MW 直接计入，不做基准容量倍数折算。
        real_caps[13] = plan[13]
        real_caps[14] = plan[14]

    return SimpleNamespace(
        e_boiler=SimpleNamespace(cap=real_caps[3], eta=0.9),
        chiller_a=SimpleNamespace(cap=real_caps[4], COP=4),
        hp_b=SimpleNamespace(cap=real_caps[9], COP=6),
        pv=SimpleNamespace(cap=real_caps[14]),
        p2g=SimpleNamespace(cap=real_caps[15], eta=0.4),
        hp_a=SimpleNamespace(cap=real_caps[8], COP=5),
        gas_turbine=SimpleNamespace(cap=real_caps[16], eta=0.7),
        chiller_b=SimpleNamespace(cap=real_caps[5], COP=5),
        gas_boiler=SimpleNamespace(cap=real_caps[7], eta=0.95),
        cchp=SimpleNamespace(cap=real_caps[17], eta_e=0.4, eta_h=0.3, eta_c=0.3),
        absorption=SimpleNamespace(cap=real_caps[6], COP=0.8),
        wind=SimpleNamespace(cap=real_caps[13]),
        storage=SimpleNamespace(
            index_map={
                "elec_s": 0,
                "elec_t": 1,
                "elec_f": 2,
                "heat_s": 3,
                "heat_t": 4,
                "heat_f": 5,
                "cool_s": 6,
                "cool_t": 7,
            }
        ),
        plan18=plan,
        real_caps=real_caps,
    )


def _load_source_constraints(attr_path: Path) -> Dict[str, np.ndarray]:
    cache_key = str(attr_path)
    cached = _CONSTRAINT_CACHE.get(cache_key)
    if cached:
        return {k: np.array(v, copy=True) if isinstance(v, np.ndarray) else v for k, v in cached.items()}

    grid = {
        "thermal_max": np.array([np.inf, np.inf, np.inf], dtype=float),
        "thermal_min": np.zeros(3),
        "gas_max": np.inf,
        "gas_min": 0.0,
        "gas_base": 0.0,
    }
    if not attr_path.exists():
        return grid

    try:
        raw = attr_path.read_text(encoding="utf-8")
        raw = _sanitize_json_text(raw)
        attr = json.loads(raw)
    except (OSError, json.JSONDecodeError):
        return grid

    sources = attr.get("top_level", {}).get("sources", [])
    for src in sources:
        src_id = src.get("id")
        params = src.get("parameters", {})
        if not src_id or not isinstance(params, dict):
            continue
        if str(src_id).startswith("thermal_gen_"):
            idx_text = str(src_id).replace("thermal_gen_", "")
            try:
                idx = int(idx_text)
            except ValueError:
                continue
            if 1 <= idx <= 3:
                grid["thermal_max"][idx - 1] = params.get("max_output_MW", params.get("max_output", grid["thermal_max"][idx - 1]))
                grid["thermal_min"][idx - 1] = params.get("min_output_MW", params.get("min_output", grid["thermal_min"][idx - 1]))
        elif src_id == "gas_source_1":
            grid["gas_max"] = params.get("max_output_MW", params.get("max_output", grid["gas_max"]))
            grid["gas_min"] = params.get("min_output_MW", params.get("min_output", grid["gas_min"]))
            grid["gas_base"] = params.get("base_injection_MW", params.get("base_injection", grid["gas_base"]))

    _CONSTRAINT_CACHE[cache_key] = {
        "thermal_max": np.array(grid["thermal_max"], copy=True),
        "thermal_min": np.array(grid["thermal_min"], copy=True),
        "gas_max": grid["gas_max"],
        "gas_min": grid["gas_min"],
        "gas_base": grid["gas_base"],
    }
    return grid


def _sanitize_json_text(raw: str) -> str:
    return raw.replace(",}", "}").replace(",]", "]")


def _parse_meos_inputs(data_dir: Path, renewable_dir: Path) -> ParsedInputs:
    cache_key = f"{data_dir}|{renewable_dir}"
    cached = _INPUT_CACHE.get(cache_key)
    if cached:
        return cached

    loads_electric = _parse_load_type(data_dir, "电", ["学生区", "教学办公区", "教工区"])
    loads_heat = _parse_load_type(data_dir, "热", ["学生区", "教学办公区", "教工区"])
    loads_cool = _parse_load_type(data_dir, "冷", ["学生区", "教学办公区", "教工区"])

    loads_total = np.column_stack([
        loads_electric[:, 0], loads_cool[:, 0], loads_heat[:, 0],
        loads_electric[:, 2], loads_cool[:, 2], loads_heat[:, 2],
        loads_electric[:, 1], loads_heat[:, 1], loads_cool[:, 1],
    ])

    prices_e, prices_g = _parse_prices(data_dir)
    pv_avail, wind_avail = _parse_renewables(renewable_dir)
    carbon_e, carbon_g = _parse_carbon_factors(data_dir)

    parsed = ParsedInputs(
        loads_total=loads_total,
        loads_electric=loads_electric,
        loads_heat=loads_heat,
        loads_cool=loads_cool,
        prices_electricity=prices_e,
        prices_gas=prices_g,
        renewable_pv_avail=pv_avail,
        renewable_wind_avail=wind_avail,
        carbon_electricity=carbon_e,
        carbon_gas=carbon_g,
    )
    _INPUT_CACHE[cache_key] = parsed
    return parsed


def _parse_load_type(data_dir: Path, load_type: str, zone_names: List[str]) -> np.ndarray:
    n_zones = len(zone_names)
    load_data = np.zeros((N_HOURS, n_zones))
    for idx, zone in enumerate(zone_names):
        csv_path = data_dir / f"负荷曲线_{load_type}_{zone}.csv"
        if csv_path.exists():
            daily = _read_daily_matrix(csv_path)
            load_data[:, idx] = _to_hourly(daily)
            continue

        summary_path = data_dir / f"汇总_负荷曲线_{load_type}.csv"
        if summary_path.exists():
            summary = _read_matrix(summary_path)
            if summary.shape[0] == 365:
                load_data[:, idx] = np.repeat(summary[:, idx], 24)
            else:
                load_data[:, idx] = _resize_to_hours(summary[:, idx], N_HOURS)
    return load_data


def _parse_prices(data_dir: Path) -> Tuple[np.ndarray, np.ndarray]:
    elec_path = data_dir / "电源价格_电.csv"
    gas_path = data_dir / "天然气价格.csv"
    summary_path = data_dir / "汇总_价格.csv"

    if elec_path.exists():
        daily = _read_daily_matrix(elec_path)
        prices_e = _to_hourly(daily)
    elif summary_path.exists():
        prices_e = _read_column(summary_path, "电价")
    else:
        prices_e = np.ones(N_HOURS) * 500

    if gas_path.exists():
        daily = _read_daily_matrix(gas_path)
        prices_g = _to_hourly(daily) / 0.01
    elif summary_path.exists():
        prices_g = _read_column(summary_path, "天然气价格") / 0.01
    else:
        prices_g = np.ones(N_HOURS) * 250

    return _resize_to_hours(prices_e, N_HOURS), _resize_to_hours(prices_g, N_HOURS)


def _parse_renewables(data_dir: Path) -> Tuple[np.ndarray, np.ndarray]:
    pv_path = data_dir / "出力曲线_光伏.csv"
    wind_path = data_dir / "出力曲线_风电.csv"
    pv_excel = data_dir / "平台光伏出力曲线.xlsx"
    wind_excel = data_dir / "平台风电出力曲线.xlsx"
    summary_path = data_dir / "汇总_出力曲线.csv"

    if pv_path.exists():
        pv = _read_daily_matrix(pv_path)
    elif pv_excel.exists():
        pv = _read_daily_matrix(pv_excel)
    elif summary_path.exists():
        pv = _read_column(summary_path, "光伏").reshape(-1, 1)
    else:
        pv = np.zeros((365, 24))

    if wind_path.exists():
        wind = _read_daily_matrix(wind_path)
    elif wind_excel.exists():
        wind = _read_daily_matrix(wind_excel)
    elif summary_path.exists():
        wind = _read_column(summary_path, "风电").reshape(-1, 1)
    else:
        wind = np.zeros((365, 24))

    pv_hourly = _resize_to_hours(_to_hourly(pv), N_HOURS)
    wind_hourly = _resize_to_hours(_to_hourly(wind), N_HOURS)

    return pv_hourly, wind_hourly


def _parse_carbon_factors(data_dir: Path) -> Tuple[np.ndarray, np.ndarray]:
    elec_path = data_dir / "购电碳排放因子.csv"
    gas_path = data_dir / "购气碳排放因子.csv"
    summary_path = data_dir / "汇总_碳排放因子.csv"

    if elec_path.exists():
        daily = _read_daily_matrix(elec_path)
        elec = _to_hourly(daily)
    elif summary_path.exists():
        elec = _read_column(summary_path, "购电碳排放因子")
    else:
        elec = np.ones(N_HOURS) * 0.6

    if gas_path.exists():
        daily = _read_daily_matrix(gas_path)
        gas = _to_hourly(daily)
    elif summary_path.exists():
        gas = _read_column(summary_path, "购气碳排放因子")
    else:
        gas = np.ones(N_HOURS) * 0.002

    return _resize_to_hours(elec, N_HOURS), _resize_to_hours(gas, N_HOURS)


def _read_daily_matrix(path: Path) -> np.ndarray:
    if path.suffix.lower() in (".xlsx", ".xls"):
        matrix = _read_excel_matrix(path)
    else:
        matrix = _read_matrix(path)
    if matrix.shape[1] >= 25:
        return matrix[:, 1:25]
    if matrix.shape[1] >= 24:
        return matrix[:, :24]
    if matrix.shape[1] == 1:
        return np.repeat(matrix, 24, axis=1)
    return matrix


def _read_matrix(path: Path) -> np.ndarray:
    rows: List[List[float]] = []
    with open(path, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        for row in reader:
            if not row:
                continue
            if _is_header_row(row):
                continue
            values = [_to_float(val) for val in row if val != ""]
            if values:
                rows.append(values)
    if not rows:
        return np.zeros((0, 0))
    max_len = max(len(r) for r in rows)
    padded = [r + [r[-1]] * (max_len - len(r)) for r in rows]
    return np.array(padded, dtype=float)


def _read_excel_matrix(path: Path) -> np.ndarray:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return np.zeros((0, 0))

    wb = load_workbook(path, data_only=True, read_only=False)
    ws = wb.active
    rows: List[List[float]] = []
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        first = row[0]
        if first is None:
            continue
        if not _is_number(str(first)):
            continue
        values = [_to_float(val) for val in row if val is not None and val != ""]
        if values:
            rows.append(values)
    if not rows:
        return np.zeros((0, 0))
    max_len = max(len(r) for r in rows)
    padded = [r + [r[-1]] * (max_len - len(r)) for r in rows]
    return np.array(padded, dtype=float)


def _read_column(path: Path, column_name: str) -> np.ndarray:
    values = []
    with open(path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if column_name not in row:
                continue
            values.append(_to_float(row.get(column_name)))
    arr = np.array(values, dtype=float)
    if arr.size == 0:
        return np.zeros(N_HOURS)
    return _resize_to_hours(arr, N_HOURS)


def _is_header_row(row: List[str]) -> bool:
    if not row:
        return False
    first = row[0].strip()
    if not first:
        return False
    return not _is_number(first)


def _to_hourly(daily: np.ndarray) -> np.ndarray:
    if daily.size == 0:
        return np.zeros(N_HOURS)
    if daily.shape[1] == 1:
        return np.repeat(daily.reshape(-1), 24)
    # Keep day-major order to match MATLAB reshape(daily', [], 1).
    return daily.reshape(-1)


def _is_number(value: str) -> bool:
    try:
        float(value)
    except ValueError:
        return False
    return True


def _to_float(value: object) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0
